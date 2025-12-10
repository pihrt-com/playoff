#!/usr/bin/env python
# -*- coding: utf-8 -*-
__author__ = 'Martin Pihrt'

APP_current       = "1.0.5"                                         # current version of the application
APP_date          = "10.12.2025-18:05"                              # current version date  
update_check_link = "https://raw.githubusercontent.com/pihrt-com/playoff/refs/heads/main/Playoff%20app/version.json"    # path to JSON with APP version

import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog, colorchooser
import math, json, os, sys
import urllib.request

# universal path to resources (works in Python and PyInstaller EXE)
if getattr(sys, 'frozen', False):
    BASE_PATH = sys._MEIPASS
else:
    BASE_PATH = os.path.abspath(os.path.dirname(__file__))

# Pillow optional
try:
    from PIL import Image, ImageDraw, ImageFont, ImageTk
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False

# try import usb_module (the new module we added). If missing, we provide a stub.
# Always add the current folder to sys.path so that the local usb_module.py is loaded.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import usb_module   # C:\Playoff\usb_module.py
    USB_AVAILABLE = True
except Exception as e:
    print("USB module error:", e)
    usb_module = None
    USB_AVAILABLE = False

# výchozí nastavení
DEFAULT_BOX_W = 180
DEFAULT_BOX_H = 30
DEFAULT_H_GAP = 200
DEFAULT_V_GAP = 10
LINE_COLOR = "#0b6bd6"
BOX_FILL = "#ffffff"
BOX_OUTLINE = "#000000"
TITLE_COLOR = "#000000"
CANVAS_BG_DEFAULT = "#ffffff"
# vítěz
WINNER_FILL = "#ffe06a"   # světle zlatá
WINNER_OUTLINE = "#cc9a00"
WINNER_TEXT_COLOR = "#000000"


# Default USB settings defaults (kept in setup file)
DEFAULT_USB_PORT = ""
DEFAULT_USB_BAUD = 9600
DEFAULT_USB_TIMEOUT = 10.0

# --- Data classes ---
class Slot:
    def __init__(self, text=""):
        self.text = text

class Match:
    def __init__(self, a=None, b=None):
        self.a = a if a else Slot()
        self.b = b if b else Slot()

class Bracket:
    def __init__(self, team_list):
        """team_list: list of initial team names (may be empty strings)"""
        self.team_count = max(0, int(len(team_list)))
        self.rounds = []  # list of lists of Match
        self.titles = []
        self._build(team_list)

    def _build(self, team_list):
        self.rounds = []
        self.titles = []
        # initial round: pair teams sequentially, allowing last single
        teams = list(team_list)
        teams = [str(t) if t is not None else "" for t in teams]

        # first round matches: ceil(n/2)
        first_matches = math.ceil(len(teams) / 2) if len(teams) > 0 else 1
        matches0 = []
        for i in range(first_matches):
            a_idx = 2 * i
            b_idx = 2 * i + 1
            a = Slot(teams[a_idx]) if a_idx < len(teams) else Slot("")
            b = Slot(teams[b_idx]) if b_idx < len(teams) else Slot("")
            matches0.append(Match(a, b))
        self.rounds.append(matches0)
        self.titles.append("Kolo 1")

        # build subsequent rounds until 1 match
        prev_matches = matches0
        r = 1
        while len(prev_matches) > 1:
            next_count = math.ceil(len(prev_matches) / 2)
            next_matches = [Match() for _ in range(next_count)]
            self.rounds.append(next_matches)
            # default title
            if next_count == 1:
                self.titles.append("Finále")
            else:
                self.titles.append(f"Kolo {r+1}")
            prev_matches = next_matches
            r += 1
        # --- Winner ---
        # last real round = final → add a round with one slot
        winner_match = Match(Slot(""), Slot(""))
        self.rounds.append([winner_match])
        self.titles.append("Vítěz")            

    def rounds_count(self):
        return len(self.rounds)

# --- App ---
class PlayoffApp:
    def __init__(self, root):
        self.root = root
        root.title("Playoff generátor by Martin Pihrt © www.pihrt.com")

        # state
        self.bracket = None
        self.current_winner = ""
        self.bg_path = None
        self.bg_image = None
        self.bg_tk = None
        self.canvas_bg = CANVAS_BG_DEFAULT
        self.line_width = 2           # normal/thick
        self.font_scale = "medium"    # small/medium/large
        self.odd_behavior = "manual"  # auto/manual/waiting
        self.lock_edit = False
        self.projector_mode = False
        self.enable_timer = True
        self.timer_value = '05:00'
        self.timer_running = False
        self.timer_blink = False
        self.current_seconds = 0
        self.timer_start_mode = "ok"   # "start" | "ok"  
        self.team_names = []              # Team naming database

        self.font_scale_var = tk.StringVar(value=self.font_scale)
        self.odd_behavior_var = tk.StringVar(value=self.odd_behavior)
        self.line_width_var = tk.IntVar(value=self.line_width)
        self.timer_menu_var = tk.BooleanVar(value=self.enable_timer)
        self.timer_start_mode_var = tk.StringVar(value=self.timer_start_mode)

        # USB manager (from usb_module)
        if USB_AVAILABLE:
            try:
                self.usb = usb_module.USBManager(self)
            except Exception:
                self.usb = None
        else:
            self.usb = None
        # store usb settings for saving into setup file
        self.usb_port = DEFAULT_USB_PORT
        self.usb_baud = DEFAULT_USB_BAUD
        self.usb_timeout = DEFAULT_USB_TIMEOUT

        # Top toolbar: left settings (calls existing menu), right Start button and status label
        toolbar = tk.Frame(root)
        toolbar.pack(side='top', fill='x', padx=4, pady=4)

        # --- Settings icon (gear wheel) ---
        try:
            settings_img_path = os.path.join(BASE_PATH, "settings.ico")

            if PIL_AVAILABLE:
                from PIL import Image, ImageTk

                img = Image.open(settings_img_path)

                # cílová výška – odpovídá Start tlačítku
                TARGET_H = 44  

                w, h = img.size
                scale = TARGET_H / h
                new_w = int(w * scale)
                new_h = int(h * scale)

                img = img.resize((new_w, new_h), Image.LANCZOS)
                settings_raw = ImageTk.PhotoImage(img)

            else:
                settings_raw = tk.PhotoImage(file=settings_img_path)

            self.settings_btn = tk.Menubutton(toolbar, image=settings_raw, relief=tk.FLAT)
            self.settings_btn.image = settings_raw


            self.settings_btn = tk.Menubutton(toolbar, image=settings_raw, relief=tk.FLAT)
            self.settings_btn.image = settings_raw  # MUSÍ být kvůli garbage collection

            self.settings_menu = tk.Menu(self.settings_btn, tearoff=0)
            self.settings_btn.config(menu=self.settings_menu)
            self.settings_btn.pack(side='left', padx=4)
            ver_txt = f"app verze: {APP_current} ({APP_date})"
            self.ver_label = tk.Label(toolbar, text=ver_txt, fg="#666666", font=("Arial", 7))
            self.ver_label.pack(side='left', padx=(10,0))

        except Exception as e:
            print("Nelze načíst settings ikonu:", e)
            # fallback
            self.settings_btn = tk.Menubutton(toolbar, text="⚙", relief=tk.RAISED)
            self.settings_menu = tk.Menu(self.settings_btn, tearoff=0)
            self.settings_btn.config(menu=self.settings_menu)
            self.settings_btn.pack(side='left')
            ver_txt = f"app verze: {APP_current} ({APP_date})"
            self.ver_label = tk.Label(toolbar, text=ver_txt, fg="#666666", font=("Arial", 7))
            self.ver_label.pack(side='left', padx=(10,0))            

        # týmy
        self.settings_menu.add_command(label='Počet týmů', command=self.ask_team_count)
        self.settings_menu.add_command(label='Generovat týmy', command=self.generate_from_entry)
        self.settings_menu.add_command(label='Pojmenování týmů', command=self.open_team_naming_dialog)
        self.settings_menu.add_command(label='Smazat obsah', command=self.reset_values)
        self.settings_menu.add_command(label='Vymazat vše', command=self.clear_all)

        # odd behavior
        odd_menu = tk.Menu(self.settings_menu, tearoff=0)
        self.settings_menu.add_cascade(label='Chování při lichém počtu vítězů', menu=odd_menu)        
        odd_menu.add_radiobutton(
            label='Automatický BYE',
            variable=self.odd_behavior_var,
            value='auto',
            command=lambda: self.set_odd_behavior('auto')
        )

        odd_menu.add_radiobutton(
            label='Ruční postup',
            variable=self.odd_behavior_var,
            value='manual',
            command=lambda: self.set_odd_behavior('manual')
        )

        odd_menu.add_radiobutton(
            label='Čekající hráč',
            variable=self.odd_behavior_var,
            value='waiting',
            command=lambda: self.set_odd_behavior('waiting')
        )

        self.settings_menu.add_separator()
        self.settings_menu.add_command(label='Načíst pozadí', command=self.load_bg_image)
        self.settings_menu.add_command(label='Smazat pozadí', command=self.remove_bg)

        # font submenu
        font_menu = tk.Menu(self.settings_menu, tearoff=0)
        self.settings_menu.add_cascade(label='Velikost písma', menu=font_menu)        
        font_menu.add_radiobutton(
            label='Malé',
            variable=self.font_scale_var,
            value='small',
            command=lambda: self.set_font_scale('small')
        )

        font_menu.add_radiobutton(
            label='Střední',
            variable=self.font_scale_var,
            value='medium',
            command=lambda: self.set_font_scale('medium')
        )

        font_menu.add_radiobutton(
            label='Velké',
            variable=self.font_scale_var,
            value='large',
            command=lambda: self.set_font_scale('large')
        )
    
        self.settings_menu.add_command(label='Barva pozadí', command=self.choose_canvas_bg)

        # line width submenu
        lw_menu = tk.Menu(self.settings_menu, tearoff=0)
        self.settings_menu.add_cascade(label='Tloušťka čar', menu=lw_menu)
        lw_menu.add_radiobutton(
            label='Normální',
            variable=self.line_width_var,
            value=2,
            command=lambda: self.set_line_width(2)
        )

        lw_menu.add_radiobutton(
            label='Silné',
            variable=self.line_width_var,
            value=4,
            command=lambda: self.set_line_width(4)
        )

        self.settings_menu.add_separator()
        self.settings_menu.add_checkbutton(label='Zamknout editaci týmů', command=self.toggle_lock_edit)
        self.settings_menu.add_command(label='Načíst ze souboru', command=self.load_setup)
        self.settings_menu.add_command(label='Uložit do souboru', command=self.save_setup)

        # timer enable
        self.settings_menu.add_separator()        
        self.timer_menu_var = tk.BooleanVar(value=self.enable_timer)
        self.settings_menu.add_checkbutton(label='Povolit odpočet', variable=self.timer_menu_var, command=self.on_toggle_timer)

        # timer mode
        timer_mode_menu = tk.Menu(self.settings_menu, tearoff=0)
        self.settings_menu.add_cascade(label='Režim spuštění odpočtu', menu=timer_mode_menu)

        timer_mode_menu.add_radiobutton(
            label='Spustit odpočet ihned (po START)',
            variable=self.timer_start_mode_var,
            value='start',
            command=lambda: self.set_timer_start_mode('start')
        )

        timer_mode_menu.add_radiobutton(
            label='Spustit odpočet až po semaforu (OK)',
            variable=self.timer_start_mode_var,
            value='ok',
            command=lambda: self.set_timer_start_mode('ok')
        )

        self.settings_menu.add_separator()
        self.settings_menu.add_command(label='USB nastavení', command=self.open_usb_dialog)

        self.settings_menu.add_separator()
        self.settings_menu.add_command(label='Export do PDF', command=self.export_pdf)
        self.settings_menu.add_command(label='Celá obrazovka ZAP/VYP', command=self.toggle_projector)
        self.settings_menu.add_command(label='Nápověda', command=self.show_help)
        
        self.settings_menu.add_command(label='Kontrola aktualizace', command=self.check_for_update)

        self.settings_btn.config(menu=self.settings_menu)
        self.settings_btn.pack(side='left')

        spacer = tk.Label(toolbar, text='')
        spacer.pack(side='left', expand=True)

        # ----- Live time and date -----
        self.datetime_var = tk.StringVar(value="")
        self.datetime_label = tk.Label(toolbar, textvariable=self.datetime_var, font=("Consolas", 22, "bold")) # 16
        self.datetime_label.pack(side='left', expand=True)

        # status label at right
        self.status_var = tk.StringVar(value='')
        self.status_label = tk.Label(toolbar, textvariable=self.status_var)
        self.status_label.pack(side='right', padx=(0,8))

        # Start button
        self.start_btn = tk.Button(
            toolbar,
            text='Start',
            command=self.on_start,
            font=("Arial", 12, "bold"),   # ← larger and bold text
            padx=15,                      # ← horizontal padding
            pady=5                        # ← vertical padding
        )

        self.start_btn.pack(side='right', padx=4)
        if not USB_AVAILABLE:
            self.start_btn.config(state='disabled')
            self.status_var.set('USB modul chybí')

        # top controls
        self.team_var = tk.StringVar(value="16")

        # canvas
        self.canvas = tk.Canvas(root, bg=self.canvas_bg)
        self.canvas.pack(fill="both", expand=True)
        self.canvas.bind('<Configure>', lambda e: self.redraw())

        # Timer overlay (canvas create_window)
        self.timer_label = tk.Label( # timer MM:SS box size
            self.root, 
            text=self.timer_value,
            font=("Consolas", 100, "bold"), 
            bg='black', 
            fg='white',
            bd=4,         # larger margin
            relief='ridge',
            padx=20,      # horizontal inner edge
            pady=10       # vertical inner edge
        ) 
        self.timer_label.bind('<Button-3>', self.on_timer_right_click)
        self.timer_window = None

        self.rect_items = {}
        self.text_items = {}
        self.title_items = {}
        self.line_items = []

        # initial bracket: empty list of N empty team slots based on team_var
        try:
            n = int(self.team_var.get())
        except Exception:
            n = 16

        self.update_datetime()

        root.bind('<Escape>', lambda e: self.exit_fullscreen())

        # try load usb settings from previous setup file if available
        # (this will be overwritten when loading a specific setup via load_setup)
        try:
            settings_path = os.path.join(os.path.expanduser('~'), '.playoff_settings.json')
            if os.path.exists(settings_path):
                with open(settings_path, 'r', encoding='utf-8') as f:
                    s = json.load(f)
                self.usb_port = s.get('usb_port', self.usb_port)
                self.usb_baud = s.get('usb_baud', self.usb_baud)
                self.usb_timeout = s.get('usb_timeout', self.usb_timeout)
                if self.usb:
                    self.usb.validate_and_set(self.usb_port, self.usb_baud, self.usb_timeout)
        except Exception:
            pass

    def check_for_update(self):
        try:
            with urllib.request.urlopen(update_check_link, timeout=3) as r:
                info = json.loads(r.read().decode('utf-8'))
        except:
            messagebox.showwarning("Aktualizace programu", "Nepodařilo se stáhnout data z Internetu.")
            return

        latest = info.get("version")
        url = info.get("url")

        if latest != APP_current:
            if messagebox.askyesno("Aktualizace programu", f"Dostupná verze {latest}. Aktualizovat?"):
                # path to the file to be updated
                if getattr(sys, 'frozen', False):
                    # runs as EXE
                    exe_path = sys.executable
                    app_dir = os.path.dirname(sys.executable)
                else:
                    # runs as .py
                    exe_path = os.path.abspath(__file__)
                    app_dir = os.path.dirname(exe_path)

                updater = os.path.join(app_dir, "updater.exe")

                import subprocess
                try:
                    subprocess.Popen([updater, exe_path, url], close_fds=True)
                except Exception as e:
                    messagebox.showerror("Updater chyba", f"Nelze spustit updater.exe:\n{e}")
                    return

                self.root.destroy()
                sys.exit(0)

        else:
            messagebox.showinfo("Aktualizace", "Máte nejnovější verzi.")
    
    def ask_team_count(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Zadat počet týmů")
        dlg.transient(self.root)
        dlg.grab_set()

        tk.Label(dlg, text="Počet týmů:").pack(padx=10, pady=6)
        ent = tk.Entry(dlg, width=10)
        ent.insert(0, self.team_var.get())
        ent.pack(padx=10, pady=4)
        ent.focus_set()

        def on_ok():
            val = ent.get()
            if not val.isdigit() or int(val) < 1:
                messagebox.showerror("Chyba", "Počet týmů musí být kladné číslo.")
                return

            if not val.isdigit() or int(val) > 24:
                messagebox.showerror("Chyba", "Počet týmů musí být < 25.")
                return                

            self.team_var.set(val)
            dlg.destroy()

            # vytvoří nový pavouk
            self.generate_bracket_with_empty(int(val))

        tk.Button(dlg, text="OK", command=on_ok).pack(pady=8)
        dlg.bind("<Return>", lambda e: on_ok())
        dlg.bind("<Escape>", lambda e: dlg.destroy())
    
    def update_datetime(self):
        import datetime
        now = datetime.datetime.now()
        self.datetime_var.set(now.strftime("%d.%m.%Y  %H:%M:%S"))
        self.root.after(1000, self.update_datetime)

    # small helper to open the existing settings menu -> we will just show the menubar cascade
    def _open_menu_settings(self):
        # open ask_team_count as a representative settings entry
        try:
            self.ask_team_count()
        except Exception:
            pass

    # --- USB / toolbar actions ---
    def on_start(self):
        # send 'start' via usb.manager asynchronously and update status label
        print("DEBUG USB PLAYOFF:",
              "usb_port=", self.usb_port,
              "usb_baud=", self.usb_baud,
              "usb_timeout=", self.usb_timeout,
              "usb obj existuje:", bool(self.usb),
              "timer_start_mode=", self.timer_start_mode)

        # At each START, first stop any running countdown.
        try:
            self.timer_running = False
            if getattr(self, 'countdown_after_id', None):
                self.root.after_cancel(self.countdown_after_id)
                self.countdown_after_id = None
        except Exception:
            pass

        self.start_btn.config(state='disabled')
        self.status_var.set('Odesílám...')
        self.status_label.config(fg='black')

        def on_result(ok, reason):
            # this callback runs in worker thread; marshal to mainloop
            def cb():
                if ok:
                    self.status_var.set('Přijato')
                    self.status_label.config(fg='green')
                    # v režimu "ok" startujeme odpočet až po "OK"
                    if self.enable_timer and self.timer_start_mode == 'ok':
                        self.start_countdown()
                else:
                    if reason == 'timeout':
                        self.status_var.set('Chyba spojení (timeout)')
                    elif isinstance(reason, str) and reason.startswith('open_error'):
                        self.status_var.set('Chyba spojení (otevření portu)')
                    elif reason == 'pyserial_missing':
                        self.status_var.set('pyserial není nainstalován')
                    else:
                        self.status_var.set('Chyba spojení (neznámá chyba)')

                    self.status_label.config(fg='red')

                    # POZOR: odpočet stopneme jen v režimu "ok".
                    # V režimu "start" musí běžet dál i při chybě spojení.
                    if self.timer_start_mode == 'ok':
                        self.timer_running = False

                self.start_btn.config(state='normal')
            try:
                self.root.after(0, cb)
            except Exception:
                pass

        # The time is ALWAYS set, but the countdown only runs in "start" mode.
        try:
            if self.enable_timer:
                self.current_seconds = self.time_to_seconds(self.timer_value)
                self.timer_label.config(text=self.seconds_to_time(self.current_seconds))

                if self.timer_start_mode == 'start':
                    # v režimu "start" se odpočet spustí vždy hned,
                    # bez ohledu na to, jestli USB vyjde nebo ne.
                    self.start_countdown()
                else:
                    # In "ok" mode, only display the time, but DO NOT START the countdown.
                    self.timer_running = False
        except Exception:
            pass

        # kick off send
        try:
            usb_available = bool(self.usb and self.usb_port)
            if usb_available:
                self.usb.validate_and_set(self.usb_port, self.usb_baud, self.usb_timeout)
                self.usb.send_start_async(on_result)
            else:
                # USB není k dispozici → nahlásíme chybu
                self.status_var.set('Chyba spojení')
                self.status_label.config(fg='red')

                # v režimu "ok" se NESMÍ nic spouštět
                if self.timer_start_mode == 'ok':
                    self.timer_running = False

            self.start_btn.config(state='normal')

        except Exception as e:
            self.status_var.set(f'Chyba: {e}')
            self.status_label.config(fg='red')
            self.start_btn.config(state='normal')
            # v režimu "ok" se při chybě NIC nespustí
            if self.timer_start_mode == 'ok':
                self.timer_running = False

    def open_team_naming_dialog(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Pojmenování týmů")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("700x500")

        # --- HLAVNÍ DATA ---
        if not hasattr(self, "team_names"):
            self.team_names = []

        # --- SCROLL KONTEJNER ---
        container = tk.Frame(dlg)
        container.pack(fill="both", expand=True, padx=10, pady=10)

        canvas = tk.Canvas(container, borderwidth=0, highlightthickness=0)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas)

        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        scroll_frame.bind_all(
            "<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        )

        # --- HLAVIČKY ---
        tk.Label(scroll_frame, text="ID týmu", font=("Arial", 10, "bold")).grid(row=0, column=0, padx=5, pady=5)
        tk.Label(scroll_frame, text="Popis týmu", font=("Arial", 10, "bold")).grid(row=0, column=1, padx=5, pady=5)
        tk.Label(scroll_frame, text="Akce", font=("Arial", 10, "bold")).grid(row=0, column=2, padx=5, pady=5)

        row_widgets = []

        def refresh_table():
            nonlocal row_widgets
            for w in row_widgets:
                for el in w:
                    el.destroy()
            row_widgets.clear()

            for i, item in enumerate(self.team_names):
                id_var = tk.StringVar(value=item.get("id", ""))
                desc_var = tk.StringVar(value=item.get("desc", ""))

                e1 = tk.Entry(scroll_frame, textvariable=id_var, width=10)
                e2 = tk.Entry(scroll_frame, textvariable=desc_var, width=40)

                e1.grid(row=i+1, column=0, padx=5, pady=2)
                e2.grid(row=i+1, column=1, padx=5, pady=2)

                def make_delete(idx):
                    return lambda: delete_row(idx)

                btn_del = tk.Button(scroll_frame, text="Smazat", command=make_delete(i))

                btn_del.grid(row=i+1, column=2, padx=5)

                def make_trace(index, var_id, var_desc):
                    def tracer(*args):
                        self.team_names[index]["id"] = var_id.get()
                        self.team_names[index]["desc"] = var_desc.get()
                    return tracer

                id_var.trace_add("write", make_trace(i, id_var, desc_var))
                desc_var.trace_add("write", make_trace(i, id_var, desc_var))

                row_widgets.append((e1, e2, btn_del))

        def add_row():
            self.team_names.append({"id": "", "desc": ""})
            refresh_table()

        def delete_row(index):
            if index < 0 or index >= len(self.team_names):
                return
            self.team_names.pop(index)
            refresh_table()

        def clear_all_rows():
            if not messagebox.askyesno("Potvrzení", "Opravdu chceš smazat všechny záznamy?"):
                return
            self.team_names.clear()
            refresh_table()

        def export_to_excel():
            if not self.team_names:
                messagebox.showinfo("Info", "Nejsou žádná data k exportu.")
                return

            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")]
            )
            if not path:
                return

            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Pojmenování týmů"

            ws.append(["ID týmu", "Popis týmu"])

            for item in self.team_names:
                ws.append([item.get("id", ""), item.get("desc", "")])

            wb.save(path)
            messagebox.showinfo("Hotovo", f"Export dokončen:\n{path}")

        def import_from_excel():
            path = filedialog.askopenfilename(
                filetypes=[("Excel", "*.xlsx")]
            )
            if not path:
                return

            from openpyxl import load_workbook

            wb = load_workbook(path)
            ws = wb.active

            self.team_names.clear()

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if not row:
                    continue
                self.team_names.append({
                    "id": str(row[0]) if row[0] is not None else "",
                    "desc": str(row[1]) if row[1] is not None else ""
                })

            refresh_table()

        # --- TLAČÍTKA ---
        btn_frame = tk.Frame(dlg)
        btn_frame.pack(fill="x", pady=8)

        tk.Button(btn_frame, text="Přidat řádek", command=add_row).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Smazat vše", command=clear_all_rows).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Import z Excelu", command=import_from_excel).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Export do Excelu", command=export_to_excel).pack(side="left", padx=5)

        def on_save():
            dlg.destroy()
            self.redraw()

        tk.Button(btn_frame, text="Uložit", command=on_save).pack(side="right", padx=5)
        tk.Button(btn_frame, text="Storno", command=dlg.destroy).pack(side="right", padx=5)

        refresh_table()
        self.root.wait_window(dlg)


    # --- settings handlers ---
    def set_odd_behavior(self, mode):
        self.odd_behavior = mode
        self.odd_behavior_var.set(mode)
        self.redraw()

    def set_font_scale(self, size):
        self.font_scale = size
        self.font_scale_var.set(size)
        self.redraw()

    def choose_canvas_bg(self):
        c = colorchooser.askcolor()[1]
        if c:
            self.canvas_bg = c
            self.canvas.config(bg=self.canvas_bg)
            self.redraw()

    def set_line_width(self, w):
        self.line_width = w
        self.line_width_var.set(w)
        self.redraw()

    def toggle_lock_edit(self):
        self.lock_edit = not self.lock_edit
        messagebox.showinfo('Info', f'Zamknutí polí bylo provedeno: {self.lock_edit}')

    def set_timer_start_mode(self, mode):
        self.timer_start_mode = mode
        self.timer_start_mode_var.set(mode)

    # --- USB dialog wrapper ---
    def open_usb_dialog(self):
        # if usb_module unavailable, show message and allow manual entry
        dlg = tk.Toplevel(self.root)
        dlg.title("USB spojení")
        dlg.transient(self.root)
        dlg.grab_set()
        w = 420; h = 320
        x = (dlg.winfo_screenwidth()-w)//2; y = (dlg.winfo_screenheight()-h)//2
        dlg.geometry(f"{w}x{h}+{x}+{y}")

        tk.Label(dlg, text="Vyber COM/USB port:").pack(anchor='w', padx=10, pady=(10,4))
        port_var = tk.StringVar(value=self.usb_port)
        if self.usb:
            ports = self.usb.list_ports()
        else:
            ports = []
        if not ports:
            # allow manual entry
            port_entry = tk.Entry(dlg, textvariable=port_var, width=30)
            port_entry.pack(padx=10, pady=4)
        else:
            # option menu + refresh
            frame = tk.Frame(dlg); frame.pack(fill='x', padx=10)
            option = tk.OptionMenu(frame, port_var, *ports)
            option.pack(side='left', fill='x', expand=True)
            def refresh_ports():
                new = self.usb.list_ports() if self.usb else []
                try:
                    option['menu'].delete(0, 'end')
                    for p in new:
                        option['menu'].add_command(label=p, command=tk._setit(port_var, p))
                except Exception:
                    pass
            tk.Button(frame, text='Obnovit', command=refresh_ports).pack(side='right', padx=6)

        tk.Label(dlg, text="Baud (rychlost) — výchozí 9600:").pack(anchor='w', padx=10, pady=(8,0))
        baud_var = tk.StringVar(value=str(self.usb_baud))
        baud_entry = tk.Entry(dlg, textvariable=baud_var, width=10)
        baud_entry.pack(padx=10, pady=4)

        tk.Label(dlg, text="Timeout (s) — výchozí 10.0:").pack(anchor='w', padx=10, pady=(8,0))
        to_var = tk.StringVar(value=str(self.usb_timeout))
        to_entry = tk.Entry(dlg, textvariable=to_var, width=10)
        to_entry.pack(padx=10, pady=4)

        def on_ok():
            self.usb_port = port_var.get().strip()
            try:
                self.usb_baud = int(baud_var.get())
            except Exception:
                self.usb_baud = DEFAULT_USB_BAUD
            try:
                self.usb_timeout = float(to_var.get())
            except Exception:
                self.usb_timeout = DEFAULT_USB_TIMEOUT
            # persist small settings file
            try:
                spath = os.path.join(os.path.expanduser('~'), '.playoff_settings.json')
                data = {}
                if os.path.exists(spath):
                    try:
                        with open(spath, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                    except Exception:
                        data = {}
                data['usb_port'] = self.usb_port
                data['usb_baud'] = self.usb_baud
                data['usb_timeout'] = self.usb_timeout
                with open(spath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
            # let usb manager validate
            try:
                if self.usb:
                    self.usb.validate_and_set(self.usb_port, self.usb_baud, self.usb_timeout)
            except Exception:
                pass
            dlg.destroy()

        def on_cancel():
            dlg.destroy()

        btnf = tk.Frame(dlg); btnf.pack(pady=8)
        tk.Button(btnf, text='OK', command=on_ok).pack(side='left', padx=6)
        tk.Button(btnf, text='Zrušit', command=on_cancel).pack(side='left', padx=6)

        dlg.bind('<Return>', lambda e: on_ok())
        dlg.bind('<Escape>', lambda e: on_cancel())
        self.root.wait_window(dlg)

    # --- bracket generation (exact number of teams) ---
    def generate_bracket_with_empty(self, n):
        # create list of n empty slots
        team_list = [""] * int(max(0, n))
        self.bracket = Bracket(team_list)
        # resolve automatic BYE if requested
        if self.odd_behavior == 'auto':
            self._auto_resolve_byes()
        self.redraw()

    def generate_from_entry(self):
        try:
            n = int(self.team_var.get())
            if n < 1:
                raise ValueError('Musí být alespoň 1 tým')
            self.generate_bracket_with_empty(n)
        except Exception:
            messagebox.showerror("Chyba", "Počet týmů není platné číslo.")

    # --- save / load setup ---
    def save_setup(self):
        if not self.bracket:
            return
        fname = filedialog.asksaveasfilename(defaultextension='.setup', filetypes=[('Playoff setup', '.setup'), ('JSON', '.json')])
        if not fname:
            return
        data = {
            'team_count': self.bracket.team_count,
            'odd_behavior': self.odd_behavior,
            'font_scale': self.font_scale,
            'canvas_bg': self.canvas_bg,
            'line_width': self.line_width,
            'bg_path': self.bg_path,
            'lock_edit': self.lock_edit,
            'titles': self.bracket.titles,
            'rounds': [],
            'winner': self.current_winner,
            'enable_timer': self.enable_timer,
            'timer_value': self.timer_value, 
            'timer_start_mode': self.timer_start_mode,                       
            # USB fields
            'usb_port': self.usb_port,
            'usb_baud': self.usb_baud,
            'usb_timeout': self.usb_timeout,
            'team_names': self.team_names,
        }
        for r in self.bracket.rounds:
            rd = []
            for m in r:
                rd.append({'a': m.a.text, 'b': m.b.text})
            data['rounds'].append(rd)
        try:
            with open(fname, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            messagebox.showinfo('Hotovo', f'Soubor byl uložen: {fname}')
        except Exception as e:
            messagebox.showerror('Chyba', str(e))

    def load_setup(self):
        fname = filedialog.askopenfilename(filetypes=[('Playoff setup', '.setup .json')])
        if not fname:
            return
        try:
            with open(fname, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror('Chyba', f'Nepodařilo se načíst soubor: {e}')
            return
        # rebuild bracket with exact count
        n = data.get('team_count', 0)
        self.generate_bracket_with_empty(n)
        self.odd_behavior = data.get('odd_behavior', 'manual')
        self.font_scale = data.get('font_scale', 'medium')
        self.canvas_bg = data.get('canvas_bg', CANVAS_BG_DEFAULT)
        self.line_width = data.get('line_width', 2)
        self.line_width_var.set(self.line_width)
        self.bg_path = data.get('bg_path')
        self.lock_edit = data.get('lock_edit', False)
        self.canvas.config(bg=self.canvas_bg)
        self.enable_timer = data.get('enable_timer', getattr(self, 'enable_timer', True))
        self.timer_value = data.get('timer_value', getattr(self, 'timer_value', '05:00'))
        self.timer_start_mode = data.get('timer_start_mode', 'start')
        self.timer_start_mode_var.set(self.timer_start_mode)
        self.team_names = data.get('team_names', [])
        try:
            # ensure timer menu var exists
            self.timer_menu_var.set(self.enable_timer)
        except Exception:
            pass
        # update timer display/visibility
        try:
            self.timer_label.config(text=self.timer_value)
            self.update_timer_visibility()
        except Exception:
            pass

        # create bracket skeleton
        self.generate_bracket_with_empty(n)
        # rebuild rounds exactly as in setup
        rounds_in = data.get('rounds', [])
        self.bracket.rounds = []  # reset completely

        for rdata in rounds_in:
            round_list = []
            for m in rdata:
                a = Slot(m.get('a', ''))
                b = Slot(m.get('b', ''))
                round_list.append(Match(a, b))
            self.bracket.rounds.append(round_list)

        # load titles
        titles = data.get('titles', [])
        for i, t in enumerate(titles):
            if i < len(self.bracket.titles):
                self.bracket.titles[i] = t
        # pokud finále obsahuje jen 1 zápas, zajistit, že se vytvořil jen 1 box
        if len(self.bracket.rounds[-1]) > 1:
            self.bracket.rounds[-1] = [self.bracket.rounds[-1][0]]
        # load rounds text
        rounds_in = data.get('rounds', [])
        for r_idx, rd in enumerate(rounds_in):
            if r_idx >= len(self.bracket.rounds):
                break
            matches = self.bracket.rounds[r_idx]
            for m_idx, mdata in enumerate(rd):
                if m_idx >= len(matches):
                    break
                matches[m_idx].a.text = mdata.get('a', '')
                matches[m_idx].b.text = mdata.get('b', '')
        # try load bg image if exists
        if self.bg_path and PIL_AVAILABLE and os.path.exists(self.bg_path):
            try:
                self.bg_image = Image.open(self.bg_path).convert('RGBA')
            except Exception:
                self.bg_image = None

        self.current_winner = data.get('winner', '')
        
        # load USB fields if present in setup file
        self.usb_port = data.get('usb_port', self.usb_port)
        self.usb_baud = data.get('usb_baud', self.usb_baud)
        self.usb_timeout = data.get('usb_timeout', self.usb_timeout)
        if self.usb:
            try:
                self.usb.validate_and_set(self.usb_port, self.usb_baud, self.usb_timeout)
            except Exception:
                pass

        self.redraw()
        messagebox.showinfo('Hotovo', 'Soubor byl načten')


    # --- Timer methods (MM:SS input, countdown, blinking) ---
    def on_toggle_timer(self):
        self.enable_timer = self.timer_menu_var.get()
        # persist small settings to user home file
        try:
            spath = os.path.join(os.path.expanduser('~'), '.playoff_settings.json')
            data = {}
            if os.path.exists(spath):
                try:
                    with open(spath, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                except Exception:
                    data = {}
            data['enable_timer'] = self.enable_timer
            data['timer_value'] = self.timer_value
            with open(spath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
        try:
            self.update_timer_visibility()
        except Exception:
            pass

    def update_timer_visibility(self):
        try:
            if getattr(self, 'timer_window', None) is None:
                return
            if self.enable_timer:
                self.canvas.itemconfigure(self.timer_window, state='normal')
            else:
                self.canvas.itemconfigure(self.timer_window, state='hidden')
        except Exception:
            try:
                if self.enable_timer:
                    self.timer_label.place(relx=0.98, rely=0.98, anchor='se')
                else:
                    self.timer_label.place_forget()
            except Exception:
                pass

    def update_timer_display(self):
        try:
            self.timer_label.config(text=self.timer_value, bg='black', fg='white')
        except Exception:
            pass

    def on_timer_right_click(self, event=None):
        class WideEntryDialog(simpledialog._QueryString):
            def body(self, master):
                super().body(master)
                self.entry.config(width=15)  # šířka pole
                return self.entry

        d = WideEntryDialog(
            title='Nastavení odpočtu',
            prompt='Zadejte čas ve formátu MM:SS',
            initialvalue=getattr(self, 'timer_value', '05:00'),
            parent=self.root
        )

        val = d.result
        if val is None:
            return
        
        if self.validate_time_format(val):
            self.timer_value = val
            self.timer_label.config(text=self.timer_value)

            # uložit nastavení
            try:
                spath = os.path.join(os.path.expanduser('~'), '.playoff_settings.json')
                data = {}
                if os.path.exists(spath):
                    with open(spath, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                data['timer_value'] = self.timer_value
                with open(spath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
        else:
            messagebox.showerror('Chyba', 'Neplatný formát času. Použij MM:SS (např. 03:30).')

    def validate_time_format(self, text):
        try:
            parts = text.split(':')
            if len(parts) != 2:
                return False
            m = int(parts[0]); s = int(parts[1])
            if m < 0 or s < 0 or s >= 60:
                return False
            return True
        except Exception:
            return False

    def time_to_seconds(self, t):
        parts = t.split(':')
        if len(parts) != 2:
            return 0
        return int(parts[0]) * 60 + int(parts[1])

    def seconds_to_time(self, sec):
        m = sec // 60
        s = sec % 60
        return f"{m:02d}:{s:02d}"

    def start_countdown(self):
        try:
            # stop blinking if active
            self.stop_blinking()
        except Exception:
            pass
        try:
            if getattr(self, 'countdown_after_id', None):
                try:
                    self.root.after_cancel(self.countdown_after_id)
                except Exception:
                    pass
                self.countdown_after_id = None
        except Exception:
            self.countdown_after_id = None
        try:
            self.current_seconds = self.time_to_seconds(self.timer_value)
            self.timer_label.config(text=self.seconds_to_time(self.current_seconds))
            self.timer_running = True
            # schedule tick
            self.countdown_after_id = self.root.after(1000, self.countdown_tick)
        except Exception:
            pass

    def countdown_tick(self):
        try:
            if not getattr(self, 'timer_running', False):
                return
            if self.current_seconds <= 0:
                self.timer_running = False
                self.start_blinking()
                return
            self.current_seconds -= 1
            try:
                self.timer_label.config(text=self.seconds_to_time(self.current_seconds))
            except Exception:
                pass
            self.countdown_after_id = self.root.after(1000, self.countdown_tick)
        except Exception:
            pass

    def start_blinking(self):
        try:
            self.timer_blink = True
            self.blink_state = False
            # kickoff blink loop
            self._blink_step()
        except Exception:
            pass

    def _blink_step(self):
        try:
            if not getattr(self, 'timer_blink', False):
                return
            self.blink_state = not getattr(self, 'blink_state', False)
            if self.blink_state:
                self.timer_label.config(bg='red', fg='white')
            else:
                self.timer_label.config(bg='white', fg='red')
            # schedule next
            self.blink_after_id = self.root.after(500, self._blink_step)
        except Exception:
            pass

    def stop_blinking(self):
        try:
            self.timer_blink = False
            if getattr(self, 'blink_after_id', None):
                try:
                    self.root.after_cancel(self.blink_after_id)
                except Exception:
                    pass
                self.blink_after_id = None
            self.timer_label.config(bg='black', fg='white')
        except Exception:
            pass


    def _auto_resolve_byes(self):
        """Automatické doplnění BYE týmů (když hraje jen jeden soupeř)."""
        rounds = self.bracket.rounds
        last_round_index = len(rounds) - 1  # poslední kolo je Vítěz

        for r_idx in range(len(rounds) - 1):
            # přeskočit kolo Vítěz
            if r_idx == last_round_index:
                continue

            matches = rounds[r_idx]

            for m_idx, match in enumerate(matches):

                a = match.a.text.strip()
                b = match.b.text.strip()

                # pokud A existuje a B je prázdné nebo naopak → automatický postup
                if (a and not b) or (b and not a):
                    winner = a if a else b

                    next_index = m_idx // 2
                    next_match = rounds[r_idx + 1][next_index]
                    side = 'a' if (m_idx % 2 == 0) else 'b'

                    if side == 'a':
                        next_match.a.text = winner
                    else:
                        next_match.b.text = winner


    # --- editing title (wide dialog) ---                       
    def edit_title(self, r_idx):
        if self.lock_edit:
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('Změnit nadpis')
        dlg.transient(self.root)
        dlg.grab_set()

        tk.Label(dlg, text='Nadpis:').pack(padx=10, pady=(10, 4))

        ent = tk.Entry(dlg, width=40)
        current = self.bracket.titles[r_idx] if r_idx < len(self.bracket.titles) else ""
        ent.insert(0, current)
        ent.pack(padx=10, pady=4)
        ent.focus_set()

        def on_ok():
            new_title = ent.get().strip()
            if r_idx < len(self.bracket.titles):
                self.bracket.titles[r_idx] = new_title
            dlg.destroy()
            self.redraw()

        def on_cancel():
            dlg.destroy()

        btns = tk.Frame(dlg)
        btns.pack(pady=10)
        tk.Button(btns, text="OK", width=10, command=on_ok).pack(side="left", padx=6)
        tk.Button(btns, text="Zrušit", width=10, command=on_cancel).pack(side="left", padx=6)

        dlg.bind("<Return>", lambda e: on_ok())
        dlg.bind("<Escape>", lambda e: on_cancel())

        self.root.wait_window(dlg)


    # --- editing UI (wide dialog) ---
    def edit_slot_dialog(self, r_idx, m_idx, side):
        if self.lock_edit:
            return
        m = self.bracket.rounds[r_idx][m_idx]
        slot = m.a if side == 'a' else m.b
        # create custom dialog
        dlg = tk.Toplevel(self.root)
        dlg.title('Zadej tým')
        dlg.transient(self.root)
        dlg.grab_set()
        w = 480
        h = 120
        x = (dlg.winfo_screenwidth() - w)//2
        y = (dlg.winfo_screenheight() - h)//2
        dlg.geometry(f'{w}x{h}+{x}+{y}')
        tk.Label(dlg, text='Zadej číslo nebo název týmu:').pack(anchor='w', padx=10, pady=(10,0))
        ent = tk.Entry(dlg, width=60)
        ent.pack(padx=10, pady=8)
        ent.insert(0, slot.text)
        ent.focus_set()
        def on_ok():
            slot.text = ent.get()
            dlg.destroy()
            self.redraw()
        def on_cancel():
            dlg.destroy()
        btnf = tk.Frame(dlg)
        btnf.pack(pady=6)
        tk.Button(btnf, text='OK', command=on_ok).pack(side='left', padx=8)
        tk.Button(btnf, text='Zrušit', command=on_cancel).pack(side='left')
        dlg.bind('<Return>', lambda e: on_ok())
        dlg.bind('<Escape>', lambda e: on_cancel())
        self.root.wait_window(dlg)

    def edit_winner_dialog(self):
        if self.lock_edit:
            return
        dlg = tk.Toplevel(self.root)
        dlg.title("Zadat vítěze")
        dlg.transient(self.root)
        dlg.grab_set()

        tk.Label(dlg, text="Jméno / číslo vítěze:").pack(padx=10, pady=(10, 4))
        ent = tk.Entry(dlg, width=40)
        ent.pack(padx=10, pady=6)
        ent.insert(0, self.current_winner)
        ent.focus_set()

        def on_ok():
            self.current_winner = ent.get().strip()
            dlg.destroy()
            self.redraw()

        def on_cancel():
            dlg.destroy()

        btn = tk.Frame(dlg)
        btn.pack(pady=10)
        tk.Button(btn, text="OK", command=on_ok).pack(side="left", padx=6)
        tk.Button(btn, text="Zrušit", command=on_cancel).pack(side="left", padx=6)

        dlg.bind("<Return>", lambda e: on_ok())
        dlg.bind("<Escape>", lambda e: on_cancel())

        self.root.wait_window(dlg)

    # --- promote (right-click) ---
    def promote(self, r_idx, m_idx, side):
        match = self.bracket.rounds[r_idx][m_idx]
        winner = match.a.text.strip() if side == 'a' else match.b.text.strip()

        if not winner:
            messagebox.showwarning('Upozornění', 'Pole je prázdné — nejdřív vyplň tým.')
            return

        # --- pokud NEJSME ve finále ---
        if r_idx + 1 < len(self.bracket.rounds):
            next_index = m_idx // 2
            next_match = self.bracket.rounds[r_idx + 1][next_index]
            next_side = 'a' if (m_idx % 2 == 0) else 'b'
            target = next_match.a if next_side == 'a' else next_match.b

            if target.text:
                if not messagebox.askyesno('Přepsat?', 'V cílovém poli už něco je. Přepsat?'):
                    return

            target.text = winner
            self.redraw()
            return

        # --- FINÁLE → zapisujeme do posledního (vítězného) kola ---
        winner_round_index = len(self.bracket.rounds) - 1   # poslední kolo je Vítěz
        winner_match = self.bracket.rounds[winner_round_index][0]

        # zapíšeme výherce do slotu A
        winner_match.a.text = winner
        winner_match.b.text = ""  # slot B je vždy prázdný

        self.redraw()
        return


    # --- reset vs clear ---
    def reset_values(self):
        # reset leaves bracket structure intact but clears all texts and titles
        if not self.bracket:
            return
        for r in self.bracket.rounds:
            for m in r:
                m.a.text = ''
                m.b.text = ''
        # keep titles as they are
        self.redraw()

    def clear_all(self):
        # clear everything: remove bracket and background; user must generate again
        # odstraníme starý pavouk
        self.bracket = None
        self.canvas.delete("all")
        self.bg_path = None
        self.bg_image = None
        self.bg_tk = None
        self.canvas_bg = CANVAS_BG_DEFAULT
        self.canvas.config(bg=self.canvas_bg)
        # reset team entry
        self.team_var.set('0')
        self.canvas.delete('all')

    # --- background functions ---
    def load_bg_image(self):
        if not PIL_AVAILABLE:
            messagebox.showerror('Chyba python Pillow', 'Pillow není nainstalován. Nainstaluj: pip install pillow')
            return
        path = filedialog.askopenfilename(title='Vyber obrázek', filetypes=[('Images','*.png;*.jpg;*.jpeg;*.bmp;*.gif')])
        if not path:
            return
        try:
            img = Image.open(path).convert('RGBA')
            self.bg_path = path
            self.bg_image = img
            self.bg_tk = None
            messagebox.showinfo('Načteno', f'Načteno: {os.path.basename(path)}')
            self.redraw()
        except Exception as e:
            messagebox.showerror('Chyba', str(e))

    def remove_bg(self):
        self.bg_path = None
        self.bg_image = None
        self.bg_tk = None
        self.redraw()
        messagebox.showinfo('Hotovo', 'Pozadí bylo odstraněno')

    # --- help ---
    def show_help(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Nápověda – Playoff generátor")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("800x600")

        frame = tk.Frame(dlg)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side="right", fill="y")

        text = tk.Text(
            frame,
            wrap="word",
            yscrollcommand=scrollbar.set,
            font=("Arial", 11)
        )
        text.pack(side="left", fill="both", expand=True)

        scrollbar.config(command=text.yview)

        help_text = (
            "NÁPOVĚDA – PLAYOFF GENERÁTOR\n"
            "Autor: Martin Pihrt © www.pihrt.com\n\n"

            "=== ZÁKLADNÍ OVLÁDÁNÍ ===\n"
            "- Počet týmů: Nastavení → Počet týmů\n"
            "- Generovat prázdné týmy: Nastavení → Generovat týmy\n"
            "- Editace týmu: Levé tlačítko myši na buňku\n"
            "- Postup do dalšího kola: Pravé tlačítko myši na tým\n"
            "- Zamknout editaci: Nastavení → Zamknout editaci týmů\n"
            "- Smazat obsah: Vymaže pouze názvy týmů\n"
            "- Vymazat vše: Odstraní pavouka i pozadí\n\n"

            "=== KOLO A TITULKY ===\n"
            "- Kliknutím na nadpis kola lze změnit jeho název\n"
            "- Poslední sloupec je vždy VÍTĚZ\n\n"

            "=== POJMENOVÁNÍ TÝMŮ ===\n"
            "Nastavení → Pojmenování týmů\n"
            "- Sloupce: ID týmu / Popis týmu\n"
            "- Neomezený počet řádků\n"
            "- Ruční editace\n"
            "- Import z Excelu\n"
            "- Export do Excelu\n"
            "- Tlačítko Smazat vše\n"
            "- Data se ukládají do souboru spolu s pavoukem\n\n"

            "=== TABULKA VPRAVO OD PAVOUKA ===\n"
            "- Automaticky se generuje z Kola 1\n"
            "- Zobrazuje pouze ID, které existují v pojmenování týmů\n"
            "- Řazení je číselně vzestupně\n"
            "- Velikost písma odpovídá nastavení velikosti pavouka\n\n"

            "=== LICHÝ POČET TÝMŮ ===\n"
            "Nastavení → Chování při lichém počtu vítězů\n"
            "- Automatický BYE\n"
            "- Ruční postup\n"
            "- Čekající hráč\n\n"

            "=== VZHLED ===\n"
            "Nastavení → Velikost písma\n"
            "- Malé / Střední / Velké\n\n"
            "Nastavení → Tloušťka čar\n"
            "- Normální / Silné\n\n"
            "Nastavení → Barva pozadí\n"
            "Nastavení → Načíst pozadí\n"
            "Nastavení → Smazat pozadí\n\n"

            "=== USB SEMAFOR ===\n"
            "Nastavení → USB nastavení\n"
            "- Výběr COM portu\n"
            "- Nastavení Baud rate (výchozí 9600)\n"
            "- Timeout odpovědi (výchozí 10 s)\n\n"
            "Tlačítko START:\n"
            "- Odešle text 'start' do zařízení\n"
            "- Očekává zpět text 'ok'\n"
            "- Stav je zobrazen v horní liště\n\n"

            "=== ODPOČET ===\n"
            "Nastavení → Povolit odpočet\n"
            "- Pod vítězem se zobrazí časovač MM:SS\n"
            "- Pravým tlačítkem myši lze změnit čas\n\n"
            "Režim spuštění odpočtu:\n"
            "- Spustit odpočet ihned (po START)\n"
            "- Spustit odpočet až po semaforu (OK)\n\n"

            "=== PDF EXPORT ===\n"
            "Nastavení → Export do PDF\n"
            "- První stránka: pavouk + vítěz + datum a čas exportu\n"
            "- Druhá stránka: tabulka pojmenování týmů\n"
            "- Automatické stránkování při dlouhé tabulce\n\n"

            "=== PROJEKTOROVÝ REŽIM ===\n"
            "Nastavení → Celá obrazovka ZAP/VYP\n"
            "- Přepne aplikaci na fullscreen\n"
            "- Klávesa ESC režim ukončí\n\n"

            "=== UKLÁDÁNÍ A NAČÍTÁNÍ ===\n"
            "Nastavení → Uložit do souboru\n"
            "- Uloží kompletní stav pavouka, pozadí, USB, odpočet i pojmenování týmů\n\n"
            "Nastavení → Načíst ze souboru\n"
            "- Obnoví veškerý uložený obsah\n\n"

            "=== AKTUALIZACE ===\n"
            "Nastavení → Kontrola aktualizace\n"
            "- Ověří novou verzi programu na internetu\n\n"

            "=== ČAS A DATUM ===\n"
            "- V horní liště se zobrazuje aktuální datum a čas v reálném čase\n"
        )

        text.insert("1.0", help_text)
        text.config(state="disabled")  # jen pro čtení

        btn = tk.Button(dlg, text="Zavřít", command=dlg.destroy)
        btn.pack(pady=8)

        dlg.bind("<Escape>", lambda e: dlg.destroy())
        self.root.wait_window(dlg)


    # --- fullscreen ---
    def toggle_projector(self):
        self.projector_mode = not getattr(self, 'projector_mode', False)
        if self.projector_mode:
            self.root.attributes('-fullscreen', True)
        else:
            self.root.attributes('-fullscreen', False)
        self.redraw()

    def exit_fullscreen(self):
        self.projector_mode = False
        self.root.attributes('-fullscreen', False)
        self.redraw()

    def build_team_lookup_from_round1(self):
        """Z Kola 1 vytvoří seřazený seznam (id, name) podle self.team_names."""
        if not self.bracket or not self.bracket.rounds:
            return []

        # 1) posbíráme všechna ID z Kola 1 (A i B)
        ids = []
        for m in self.bracket.rounds[0]:
            a = m.a.text.strip()
            b = m.b.text.strip()
            if a:
                ids.append(a)
            if b:
                ids.append(b)

        # 2) databáze pojmenování → dict pro rychlé hledání
        lookup = {str(x.get("id", "")).strip(): x.get("desc", "") for x in self.team_names}

        # 3) vytvoříme pouze páry, které existují v pojmenování
        result = []
        for tid in ids:
            if tid in lookup:
                result.append((tid, lookup[tid]))

        # 4) číselné řazení podle ID
        def sort_key(x):
            try:
                return int(x[0])
            except Exception:
                return 10**9

        result.sort(key=sort_key)
        return result

    # --- redraw ---
    def redraw(self):
        self.canvas.delete('all')
        self.rect_items.clear()
        self.text_items.clear()
        self.title_items.clear()
        self.line_items.clear()

        if not self.bracket:
            return

        width = self.root.winfo_screenwidth() if self.projector_mode else max(800, self.canvas.winfo_width())
        height = self.root.winfo_screenheight() if self.projector_mode else max(600, self.canvas.winfo_height())
        self.canvas.config(width=width, height=height, bg=self.canvas_bg)

        rounds = self.bracket.rounds
        rounds_count = len(rounds)
        last_round = rounds_count - 1
        final_real_round = last_round - 1

        # --- font sizes ---
        if self.font_scale == 'large':
            cell_font_size = 28
        elif self.font_scale == 'small':
            cell_font_size = 10
        else:
            cell_font_size = 14

        from tkinter import font as tkfont
        cell_font = tkfont.Font(family="Arial", size=cell_font_size)

        title_font_size = 33 if self.projector_mode or self.font_scale == 'large' else (20 if self.font_scale == 'medium' else 12)
        title_font = ("Arial", title_font_size, "bold")

        # --- column widths ---
        col_widths = []
        padding = 20
        min_w = 60
        max_w = 220

        for r_idx, matches in enumerate(rounds):
            if r_idx == last_round:
                winner_match = rounds[last_round][0]
                winner_text = winner_match.a.text.strip() or winner_match.b.text.strip() or self.current_winner
                text_px = cell_font.measure(winner_text)
                col_widths.append(max(min_w, min(max_w, text_px + padding)))
                continue

            max_px = 0
            for m in matches:
                max_px = max(max_px, cell_font.measure(m.a.text.strip()))
                max_px = max(max_px, cell_font.measure(m.b.text.strip()))
            col_widths.append(max(min_w, min(max_w, max_px + padding)))

        margin_x = 40
        margin_y = 40
        h_gap = 180
        v_gap = 10

        box_h = 30
        per_match_h = box_h * 2 + 8
        initial_matches = len(rounds[0])

        total_h = initial_matches * per_match_h + (initial_matches - 1) * v_gap + 2 * margin_y
        if total_h > height:
            scale = height / total_h
            scale = max(0.25, scale)
            box_h = int(box_h * scale)
            v_gap = max(3, int(v_gap * scale))
            per_match_h = box_h * 2 + 8

        col_x_positions = []
        x = margin_x
        for w in col_widths:
            col_x_positions.append(x)
            x += w + h_gap

        final_centers = []

        for r_idx, matches in enumerate(rounds):
            col_x = col_x_positions[r_idx]
            box_w = col_widths[r_idx]

            title_id = self.canvas.create_text(
                col_x + box_w/2,
                margin_y - 10,
                text=self.bracket.titles[r_idx],
                font=title_font,
                fill=TITLE_COLOR
            )
            self.canvas.tag_bind(title_id, '<Button-1>', lambda e, rr=r_idx: self.edit_title(rr))

            matches_count = len(matches)
            total_col_h = matches_count * (box_h * 2 + 8) + (matches_count - 1) * v_gap
            start_y = max(margin_y + title_font_size + 10,
                          (height - total_col_h) // 2)

            for m_idx, match in enumerate(matches):

                if r_idx == last_round and len(matches) == 1:
                    continue

                top_y = start_y + m_idx * ((box_h * 2 + 8) + v_gap)

                match_x1 = col_x
                match_y1 = top_y
                match_x2 = match_x1 + box_w
                match_y2 = match_y1 + box_h * 2 + 8

                a_x1, a_y1, a_x2, a_y2 = match_x1, match_y1, match_x2, match_y1 + box_h
                b_x1, b_y1, b_x2, b_y2 = match_x1, a_y2 + 8, match_x2, match_y2

                rect_match = self.canvas.create_rectangle(
                    match_x1, match_y1, match_x2, match_y2,
                    fill=BOX_FILL, outline=BOX_OUTLINE, width=self.line_width
                )

                self.canvas.create_line(
                    a_x1, a_y2 + 4,
                    a_x2, a_y2 + 4,
                    fill=BOX_OUTLINE,
                    width=max(1, self.line_width - 1)
                )

                is_winner_a = (match.a.text.strip() == self.current_winner.strip())
                is_winner_b = (match.b.text.strip() == self.current_winner.strip())

                if is_winner_a:
                    self.canvas.create_rectangle(a_x1, a_y1, a_x2, a_y2, fill=WINNER_FILL, outline="", width=0)
                if is_winner_b:
                    self.canvas.create_rectangle(b_x1, b_y1, b_x2, b_y2, fill=WINNER_FILL, outline="", width=0)

                txt_a = self.canvas.create_text((a_x1+a_x2)/2, (a_y1+a_y2)/2,
                                                text=match.a.text, font=("Arial", cell_font_size))
                txt_b = self.canvas.create_text((b_x1+b_x2)/2, (b_y1+b_y2)/2,
                                                text=match.b.text, font=("Arial", cell_font_size))

                # HITBOXY MUSÍ BÝT POSLEDNÍ
                hit_a = self.canvas.create_rectangle(a_x1, a_y1, a_x2, a_y2, fill="", outline="")
                hit_b = self.canvas.create_rectangle(b_x1, b_y1, b_x2, b_y2, fill="", outline="")

                if not self.lock_edit:
                    self.canvas.tag_bind(hit_a, '<Button-1>', lambda e, rr=r_idx, mm=m_idx: self.edit_slot_dialog(rr, mm, 'a'))
                    self.canvas.tag_bind(hit_b, '<Button-1>', lambda e, rr=r_idx, mm=m_idx: self.edit_slot_dialog(rr, mm, 'b'))
                    self.canvas.tag_bind(hit_a, '<Button-3>', lambda e, rr=r_idx, mm=m_idx: self.promote(rr, mm, 'a'))
                    self.canvas.tag_bind(hit_b, '<Button-3>', lambda e, rr=r_idx, mm=m_idx: self.promote(rr, mm, 'b'))

                if r_idx + 1 < rounds_count:
                    cx = a_x2
                    cy = (a_y1 + b_y2) / 2

                    next_col_x = col_x_positions[r_idx + 1]
                    next_idx = m_idx // 2
                    nm = len(rounds[r_idx + 1])
                    total_h_next = nm * (box_h * 2 + 8) + (nm - 1) * v_gap
                    start_y_next = max(margin_y + title_font_size + 10,
                                       (height - total_h_next) // 2)
                    ny = start_y_next + next_idx * ((box_h * 2 + 8) + v_gap) + box_h

                    self.canvas.create_line(cx, cy, next_col_x, ny,
                                            width=self.line_width, fill=LINE_COLOR, smooth=True)

                if r_idx == final_real_round:
                    final_centers.append(((a_x2), (a_y1 + b_y2) / 2))

        # --- Winner box ---
        winner_match = rounds[last_round][0]
        winner_text = winner_match.a.text.strip() or winner_match.b.text.strip() or self.current_winner

        text_width = cell_font.measure(winner_text) if winner_text else 0
        win_w = max(text_width + 60, col_widths[last_round] + 40)
        win_h = box_h * 2

        col_x = col_x_positions[last_round]
        win_x_center = col_x + col_widths[last_round] / 2

        win_x1 = win_x_center - win_w/2
        win_x2 = win_x_center + win_w/2
        win_y1 = (height // 2) - win_h//2
        win_y2 = win_y1 + win_h

        self.canvas.create_rectangle(
            win_x1, win_y1, win_x2, win_y2,
            fill=BOX_FILL, outline=BOX_OUTLINE, width=self.line_width+1
        )

        self.canvas.create_text(
            (win_x1+win_x2)/2, (win_y1+win_y2)/2,
            text=winner_text,
            font=("Arial", cell_font_size*2, "bold")
        )


        # one line between the final and the winner = we will use the AVERAGE of both centers
        if final_centers:
            avg_x = sum(p[0] for p in final_centers) / len(final_centers)
            avg_y = sum(p[1] for p in final_centers) / len(final_centers)
            self.line_items.append((avg_x, avg_y, win_x1, (win_y1 + win_y2) / 2, self.line_width, LINE_COLOR))

        # === TABULKA POJMENOVÁNÍ TÝMŮ (Z KOLA 1, VPRAVO OD PAVOUKA) ===
        try:
            table_data = self.build_team_lookup_from_round1()
            if table_data:

                # --- geometrie prostoru mezi pavoukem a vítězem ---
                table_x1 = win_x2 + 20
                table_x2 = width - 20

                # pokud není prostor, tabulku nevykreslujeme
                if table_x2 - table_x1 > 200:

                    table_y = margin_y + 40

                    # --- řízení velikosti písma tabulky vůči pavouku ---
                    table_font_size = min(cell_font_size, 18)     # <<< HORNÍ OMEZENÍ
                    row_h = max(26, table_font_size + 10)         # <<< DYNAMICKÁ VÝŠKA ŘÁDKU

                    col1_w = 90
                    col2_w = table_x2 - table_x1 - col1_w - 10

                    # --- HLAVIČKA ---
                    self.canvas.create_rectangle(
                        table_x1, table_y,
                        table_x2, table_y + row_h,
                        fill="#dddddd", outline=BOX_OUTLINE, width=1
                    )

                    self.canvas.create_text(
                        table_x1 + col1_w/2, table_y + row_h/2,
                        text="ID", font=("Arial", min(16, table_font_size), "bold")
                    )

                    col_title = self.bracket.titles[0] if self.bracket.titles else "Popis"
                    self.canvas.create_text(
                        table_x1 + col1_w + col2_w/2 + 5, table_y + row_h/2,
                        text=col_title, font=("Arial", min(16, table_font_size), "bold")
                    )

                    y = table_y + row_h

                    # --- ŘÁDKY ---
                    for tid, name in table_data:
                        self.canvas.create_rectangle(
                            table_x1, y,
                            table_x2, y + row_h,
                            fill="#ffffff", outline=BOX_OUTLINE, width=1
                        )

                        self.canvas.create_text(
                            table_x1 + col1_w/2, y + row_h/2,
                            text=str(tid), font=("Arial", table_font_size)
                        )

                        self.canvas.create_text(
                            table_x1 + col1_w + col2_w/2 + 5, y + row_h/2,
                            text=name, font=("Arial", table_font_size), anchor="center"
                        )

                        y += row_h

        except Exception as e:
            print("TEAM TABLE ERROR:", e)

        # --- TIMER OVERLAY (ALWAYS RE-CREATE AFTER canvas.delete('all')) ---
        try:
            if self.enable_timer:
                # pozice pod boxem vítěze
                tx = win_x2
                ty = win_y2 + 30

                # Timer window je vždy ZNOVU vytvořen po delete('all')
                # → Therefore, the old ID is invalid and must be regenerated.
                self.timer_window = self.canvas.create_window(
                    tx, ty,
                    window=self.timer_label,
                    anchor='ne'
                )

                # The timer must be visible.
                self.canvas.itemconfigure(self.timer_window, state='normal')

            else:
                # timer off → hide widget
                if self.timer_window:
                    try:
                        self.canvas.itemconfigure(self.timer_window, state='hidden')
                    except:
                        pass

        except Exception as e:
            print("TIMER ERROR:", e)


    # --- PDF export ---
    def export_pdf(self):
        try:
            from reportlab.pdfgen import canvas as rl_canvas
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.colors import HexColor, black
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            messagebox.showerror(
                "Chyba",
                "Pro vektorový PDF export je potřeba knihovna reportlab:\n\npip install reportlab"
            )
            return

        if not self.bracket:
            messagebox.showerror("Chyba", "Nejprve vytvoř pavouka.")
            return

        # --- Font ---
        try:
            pdfmetrics.registerFont(TTFont("DejaVu", "DejaVuSans.ttf"))
        except Exception:
            messagebox.showerror(
                "Chyba",
                "Font DejaVuSans.ttf nebyl nalezen.\nUlož jej do stejné složky jako tento program."
            )
            return

        fname = filedialog.asksaveasfilename(defaultextension=".pdf",
                                             filetypes=[("PDF", "*.pdf")])
        if not fname:
            return

        # --- Geometry as redraw() ---
        rounds = self.bracket.rounds
        rounds_count = len(rounds)
        last_round = rounds_count - 1
        final_real_round = last_round - 1

        # canvas size
        canvas_w = self.root.winfo_screenwidth() if self.projector_mode else max(800, self.canvas.winfo_width())
        canvas_h = self.root.winfo_screenheight() if self.projector_mode else max(600, self.canvas.winfo_height())

        # font size
        if self.font_scale == 'large':
            cell_font_size = 28
        elif self.font_scale == 'small':
            cell_font_size = 10
        else:
            cell_font_size = 14

        from tkinter import font as tkfont
        cell_font = tkfont.Font(family="Arial", size=cell_font_size)

        title_font_size = (
            33 if self.projector_mode or self.font_scale == 'large'
            else 20 if self.font_scale == 'medium'
            else 12
        )

        # --- column widths ---
        col_widths = []
        padding = 20
        min_w = 60
        max_w = 220

        for r_idx, matches in enumerate(rounds):
            if r_idx == last_round:
                winner_match = rounds[last_round][0]
                winner_text = winner_match.a.text.strip() or winner_match.b.text.strip() or self.current_winner
                text_px = cell_font.measure(winner_text)
                col_widths.append(max(min_w, min(max_w, text_px + padding)))
                continue

            max_px = 0
            for m in matches:
                max_px = max(max_px, cell_font.measure(m.a.text.strip()))
                max_px = max(max_px, cell_font.measure(m.b.text.strip()))
            col_widths.append(max(min_w, min(max_w, max_px + padding)))

        # --- indents and spaces ---
        margin_x = 40
        margin_y = 40
        h_gap = 180
        v_gap = 10

        # --- wheel height ---
        box_h = 30
        per_match_h = box_h * 2 + 8
        initial_matches = len(rounds[0])

        total_h = initial_matches * per_match_h + (initial_matches - 1) * v_gap + 2 * margin_y
        if total_h > canvas_h:
            scale_local = canvas_h / total_h
            scale_local = max(0.25, scale_local)
            box_h = int(box_h * scale_local)
            v_gap = max(3, int(v_gap * scale_local))
            per_match_h = box_h * 2 + 8

        # --- X column positions ---
        col_x_positions = []
        x = margin_x
        for w in col_widths:
            col_x_positions.append(x)
            x += w + h_gap

        # --- Geometry for PDF ---
        rects = []
        lines = []
        titles = []

        # We will save the ONLY point of the final
        final_center_x = None
        final_center_y = None

        # --- Wheel scanning ---
        for r_idx, matches in enumerate(rounds):
            col_x = col_x_positions[r_idx]
            box_w = col_widths[r_idx]

            # titles
            titles.append((col_x + box_w / 2, margin_y - 10,
                           self.bracket.titles[r_idx], title_font_size))

            matches_count = len(matches)
            total_col_h = matches_count * (box_h * 2 + 8) + (matches_count - 1) * v_gap
            start_y = max(margin_y + title_font_size + 10,
                          (canvas_h - total_col_h) // 2)

            for m_idx, match in enumerate(matches):

                # skip the WINNERS round (to be drawn later)
                if r_idx == last_round and len(matches) == 1:
                    continue

                top_y = start_y + m_idx * ((box_h * 2 + 8) + v_gap)

                # A box
                a_x1 = col_x
                a_y1 = top_y
                a_x2 = a_x1 + box_w
                a_y2 = a_y1 + box_h

                # B box
                b_x1 = col_x
                b_y1 = a_y2 + 8
                b_x2 = b_x1 + box_w
                b_y2 = b_y1 + box_h

                # winner/highlight
                is_winner_a = (match.a.text.strip() == self.current_winner.strip())
                is_winner_b = (match.b.text.strip() == self.current_winner.strip())

                fill_a = WINNER_FILL if is_winner_a else BOX_FILL
                fill_b = WINNER_FILL if is_winner_b else BOX_FILL
                outline_a = WINNER_OUTLINE if is_winner_a else BOX_OUTLINE
                outline_b = WINNER_OUTLINE if is_winner_b else BOX_OUTLINE

                rects.append((a_x1, a_y1, a_x2, a_y2,
                              self.line_width, fill_a, outline_a,
                              match.a.text, cell_font_size, False))
                rects.append((b_x1, b_y1, b_x2, b_y2,
                              self.line_width, fill_b, outline_b,
                              match.b.text, cell_font_size, False))

                # connection to the next round
                if (r_idx + 1 < rounds_count) and (r_idx != final_real_round):                    
                    cx = a_x2
                    cy = (a_y1 + b_y2) / 2

                    next_col_x = col_x_positions[r_idx + 1]
                    next_idx = m_idx // 2
                    nm = len(rounds[r_idx + 1])
                    total_h_next = nm * (box_h * 2 + 8) + (nm - 1) * v_gap
                    start_y_next = max(margin_y + title_font_size + 10,
                                       (canvas_h - total_h_next) // 2)
                    ny = start_y_next + next_idx * ((box_h * 2 + 8) + v_gap) + box_h

                    lines.append((cx, cy, next_col_x, ny,
                                  self.line_width, LINE_COLOR))

                # --- THE ONLY center of the finals ---
                if r_idx == final_real_round and final_center_x is None:
                    final_center_x = a_x2
                    final_center_y = (a_y1 + b_y2) / 2

        # --- Drawing of the winner ---
        winner_match = rounds[last_round][0]
        winner_text = (winner_match.a.text.strip()
                       or winner_match.b.text.strip()
                       or self.current_winner)

        text_width = cell_font.measure(winner_text)
        win_w = max(text_width + 60, col_widths[last_round] + 40)
        win_h = box_h * 2

        col_x = col_x_positions[last_round]
        win_x_center = col_x + col_widths[last_round] / 2

        win_x1 = win_x_center - win_w / 2
        win_x2 = win_x_center + win_w / 2
        win_y1 = (canvas_h // 2) - win_h // 2
        win_y2 = win_y1 + win_h

        rects.append((win_x1, win_y1, win_x2, win_y2,
                      self.line_width + 1, BOX_FILL, BOX_OUTLINE,
                      winner_text, cell_font_size * 2, True))

        # --- ONE line from the final → winner ---
        if final_center_x is not None:
            lines.append((final_center_x, final_center_y,
                          win_x1, (win_y1 + win_y2) / 2,
                          self.line_width, LINE_COLOR))

        # --- Bounding box calculation ---
        minx, miny = float('inf'), float('inf')
        maxx, maxy = float('-inf'), float('-inf')

        for x1, y1, x2, y2, *_ in rects:
            minx = min(minx, x1)
            miny = min(miny, y1)
            maxx = max(maxx, x2)
            maxy = max(maxy, y2)

        for x1, y1, x2, y2, *_ in lines:
            minx = min(minx, x1, x2)
            miny = min(miny, y1, y2)
            maxx = max(maxx, x1, x2)
            maxy = max(maxy, y1, y2)

        for tx, ty, *_ in titles:
            minx = min(minx, tx)
            miny = min(miny, ty)
            maxx = max(maxx, tx)
            maxy = max(maxy, ty)

        # --- Reserve for printing ---
        maxy += 60

        content_w = maxx - minx
        content_h = maxy - miny

        # --- PDF page ---
        page_w, page_h = landscape(A4)
        margin_pdf = 20

        scale_x = (page_w - 2 * margin_pdf) / content_w
        scale_y = (page_h - 2 * margin_pdf) / content_h
        scale = min(scale_x, scale_y)

        draw_w = content_w * scale
        draw_h = content_h * scale

        off_x = (page_w - draw_w) / 2 - minx * scale
        off_y = (page_h - draw_h) / 2 - miny * scale

        def TX(x): return off_x + x * scale
        def TY(y): return off_y + (content_h * scale) - (y - miny) * scale

        # --- PDF ---
        pdf = rl_canvas.Canvas(fname, pagesize=(page_w, page_h))

        # --- RECTANGLES & TEXT ---
        for x1, y1, x2, y2, lw, fill, outline, text, fsize, bold in rects:
            rw = (x2 - x1) * scale
            rh = (y2 - y1) * scale
            rx = TX(x1)
            ry_top = TY(y1)
            ry = ry_top - rh

            pdf.setStrokeColor(HexColor(outline))
            pdf.setFillColor(HexColor(fill))
            pdf.setLineWidth(max(0.5, lw * scale))
            pdf.rect(rx, ry, rw, rh, stroke=1, fill=1)

            # text
            font_pt = max(6, int(fsize * scale))
            baseline_shift = font_pt * 0.35

            txt_cx = rx + rw / 2
            txt_cy = ry + rh / 2

            pdf.setFont("DejaVu", font_pt)
            pdf.setFillColor(black)

            pdf.drawCentredString(txt_cx, txt_cy - baseline_shift, text)

        # --- TITLES ---
        for tx, ty, ttext, tsize in titles:
            font_pt = max(7, int(tsize * scale))
            baseline_shift = font_pt * 0.35
            px = TX(tx)
            py_top = TY(ty)
            py = py_top - baseline_shift

            pdf.setFont("DejaVu", font_pt)
            pdf.setFillColor(black)
            pdf.drawCentredString(px, py, ttext)

        # --- LINES ---
        for x1, y1, x2, y2, lw, color in lines:
            pdf.setStrokeColor(HexColor(color))
            pdf.setLineWidth(max(0.5, lw * scale))
            pdf.line(TX(x1), TY(y1), TX(x2), TY(y2))

        # --- DATUM A ČAS EXPORTU ---
        import datetime
        dt = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")

        pdf.setFont("DejaVu", 10)
        pdf.setFillColor(black)
        pdf.drawRightString(
            page_w - 20,
            20,
            f"Export: {dt}"
        )

        # === TABULKA POJMENOVÁNÍ TÝMŮ NA SAMOSTATNÉ STRÁNCE PDF ===
        try:
            table_data = self.build_team_lookup_from_round1()
            if table_data:
                pdf.showPage()   # ⬅⬅⬅ DRUHÁ STRANA

                page_w, page_h = landscape(A4)
                margin = 40

                col1_w = 100
                col2_w = page_w - col1_w - 2 * margin
                row_h = max(26, min(cell_font_size, 18) + 10)

                x1 = margin
                y = page_h - margin

                # --- HLAVIČKA ---
                pdf.setFillColor(HexColor("#dddddd"))
                pdf.rect(x1, y - row_h, col1_w, row_h, stroke=1, fill=1)
                pdf.rect(x1 + col1_w, y - row_h, col2_w, row_h, stroke=1, fill=1)

                pdf.setFont("DejaVu", 12)
                pdf.setFillColor(black)
                pdf.drawCentredString(x1 + col1_w / 2, y - row_h / 2 - 4, "ID")

                col_title = self.bracket.titles[0] if self.bracket.titles else "Kolo 1"
                pdf.drawCentredString(x1 + col1_w + col2_w / 2, y - row_h / 2 - 4, col_title)

                y -= row_h
                pdf.setFont("DejaVu", 11)

                for tid, name in table_data:

                    # --- STRÁNKOVÁNÍ ---
                    if y - row_h < margin:
                        pdf.showPage()
                        y = page_h - margin

                        pdf.setFillColor(HexColor("#dddddd"))
                        pdf.rect(x1, y - row_h, col1_w, row_h, stroke=1, fill=1)
                        pdf.rect(x1 + col1_w, y - row_h, col2_w, row_h, stroke=1, fill=1)

                        pdf.setFont("DejaVu", 12)
                        pdf.setFillColor(black)
                        pdf.drawCentredString(x1 + col1_w / 2, y - row_h / 2 - 4, "ID")
                        pdf.drawCentredString(x1 + col1_w + col2_w / 2, y - row_h / 2 - 4, col_title)

                        y -= row_h
                        pdf.setFont("DejaVu", 11)

                    pdf.setFillColor(HexColor("#ffffff"))
                    pdf.rect(x1, y - row_h, col1_w, row_h, stroke=1, fill=1)
                    pdf.rect(x1 + col1_w, y - row_h, col2_w, row_h, stroke=1, fill=1)

                    pdf.setFillColor(black)
                    pdf.drawCentredString(x1 + col1_w / 2, y - row_h / 2 - 4, str(tid))
                    pdf.drawString(x1 + col1_w + 6, y - row_h / 2 - 4, name)

                    y -= row_h

                pdf.setFont("DejaVu", 10)
                pdf.setFillColor(black)
                pdf.drawRightString(
                    page_w - 20,
                    20,
                    f"Export: {dt}"
                )

        except Exception as e:
            print("PDF TEAM TABLE ERROR:", e)

        # --- Save PDF ---
        try:
            pdf.save()
            messagebox.showinfo("Hotovo", f"PDF uložen: {fname}")
        except Exception as e:
            messagebox.showerror("Chyba", str(e))

# --- run ---
if __name__ == '__main__':
    root = tk.Tk()

    # --- FIX icon for PyInstaller ONEFILE ---
    try:
        if getattr(sys, 'frozen', False):
            base = sys._MEIPASS
        else:
            base = os.path.dirname(os.path.abspath(__file__))

        icon_path = os.path.join(base, "playoff.ico")
        root.iconbitmap(icon_path)

    except Exception as e:
        print("Icon load error:", e)

    app = PlayoffApp(root)
    root.geometry('1920x1080')
    root.mainloop()